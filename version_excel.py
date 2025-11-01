import json 
import re
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

def read_config(filename):
    config = {}
    with open(filename, "r") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):  # ignorer commentaires et lignes vides
                key, value = line.split("=", 1)
                config[key.strip()] = value.strip()
    return config

# Charger le fichier de config
config = read_config("config.txt")

# Paramètres
input_file = config.get("input_file", "input/test.txt")
output_file = config.get("output_file", "output/test_partoche.xlsx")

# Mapping CSS → hex pour openpyxl
css_to_hex = {
    "purple": "800080",
    "orange": "FFA500",
    "blue": "0000FF",
    "green": "008000",
    "red": "FF0000",
    "yellow": "FFD700",
    "gold": "FFD700"
}

# Charger le mapping des couleurs depuis conversion.json
with open("conversion.json", "r") as f:
    color_map = json.load(f)

# Charger le fichier input
with open(input_file, "r") as f:
    input_lines = [line.strip() for line in f.readlines()]

# Colonnes de A à G pour la grille
columns = ["A", "B", "C", "D", "E", "F", "G"]

def clean_note(note):
    """
    Retire le chiffre de l'octave et garde tous les caractères spéciaux autour de la lettre.
    """
    match = re.search(r"[A-G]", note)
    if not match:
        return note
    letter = match.group(0)
    prefix_match = re.match(r"[^A-G\d]*", note)
    prefix = prefix_match.group(0) if prefix_match else ""
    suffix_match = re.search(r"[^0-9]*$", note)
    suffix = suffix_match.group(0) if suffix_match else ""
    return f"{prefix}{letter}{suffix}"

# Créer le workbook Excel
wb = Workbook()
ws = wb.active
ws.title = "PianoRoll"

num_grids = 6  # nombre de grilles
start_col = 1  # colonne de départ dans Excel

# Ajouter le header avec les lettres colorées
for grid_index in range(num_grids):
    grid_color = color_map.get(str(grid_index + 1), "000000")
    for col_idx, col_name in enumerate(columns):
        excel_col = start_col + grid_index*len(columns) + col_idx
        cell = ws.cell(row=1, column=excel_col, value=col_name)
        # Appliquer la couleur uniquement à la lettre
        hex_color = css_to_hex.get(grid_color.lower(), "000000")
        cell.font = Font(color=hex_color, bold=True)

# Remplir les lignes de notes (vides si pas de note)
for row_idx, line in enumerate(input_lines, start=2):
    notes_in_line = line.split()
    for grid_index in range(num_grids):
        grid_color = color_map.get(str(grid_index + 1), "000000")
        for col_idx, note_col in enumerate(columns):
            cell_value = ""
            for note in notes_in_line:
                letter_match = re.search(r"[A-G]", note)
                if letter_match:
                    letter = letter_match.group(0)
                    num_match = re.search(r'\d+', note)
                    number = int(num_match.group()) if num_match else 1
                    if number == grid_index + 1 and letter == note_col:
                        cell_value = clean_note(note)
                        break
            # Calculer la colonne Excel
            excel_col = start_col + grid_index*len(columns) + col_idx
            cell = ws.cell(row=row_idx, column=excel_col, value=cell_value)
            # Pas de remplissage de couleur
            cell.font = Font(bold=True, color=css_to_hex.get(grid_color.lower(), "000000") if cell_value else "000000")

# Vérifier si le fichier existe
if os.path.exists(output_file):
    # Récupérer la date et l'heure actuelles
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Créer un nouveau nom avec timestamp
    base, ext = os.path.splitext(output_file)
    new_name = f"{base}_{timestamp}{ext}"
    os.rename(output_file, new_name)
    print(f"Fichier existant renommé en {new_name}")

# Sauvegarder le nouveau fichier Excel
wb.save(output_file)
print(f"Fichier {output_file} généré avec succès !")